from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from datetime import datetime
from datetime import time
import re
import io
import sys
import os.path


excelFile = 'ММГ с 21.03.2022 по 27.03.2022.xlsx'
"""excelFile = None

if sys.argv[1].strip() == '':
    print('')
    print('Необходимо указать путь до файла первым аргуметом!')
    exit()
else:
    if os.path.isfile(sys.argv[1].strip()) and not os.path.isdir(sys.argv[1].strip()):
        excelFile = sys.argv[1].strip()
    else:
        print('')
        print("Указан неверный путь к файлу %s" % sys.argv[1].strip())
        exit()"""


# Путь до файла с данными
excelFileOut = '2022.03.29_' + excelFile
rowStart = 2  # Первая строка данных (после заголовка)

# Уже заполненные поля
columnDoc = 'P'  # Врач
columnDocExpert = 'Q'  # Эксперт
columnResolution = 'L'  # Заключение
columnDescription = 'K'  # Описание
columnOrg = 'A'  # Организация
columnDate = 'D'  # Дата исследования
columnUid = 'C'  # UID исследования
columnDateEdit = 'E'  # Дата создания записи

# Необходимо создать пустые столбики для заполнения данных
columnPyResolution = 'Z'  # Исправленное заключение
columnPyResolutionCheck = 'AA'  # Исправленное заключение для поска
columnPyRads = 'W'  # RADS
columnPgmi = 'X'  # RADS
columnError = 'Y'  # Error
columnPyResult = 'AB'  # Комметарии
columnErrorComment = 'AC'  # Error
columnKicker = 'AF'  # Org + Date

columnVerboseRads = 'AD'
columnVerbosePgmi = 'AE'

regexpFindPgmi = \
    re.compile(r"PGMI\s*\:?[\s]*.?([PGMIРМ])", re.IGNORECASE | re.MULTILINE)

# Регулярное выражение для поиска всех вариантов "радса"
regexpFindRads = \
    re.compile(r"(?:BI(?:[-\s]{0,10})?)?(R?ADS)[^0123456IV\n]{0,15}([0123456]|[IV]{1,4})([^0123456IV]|$)", re.IGNORECASE)
replaceRads = r"[RADS\2]\3"

regexpFindMol = \
    re.compile(r"((?:Прав\w\w|Лев\w\w)\s+молочн\w\w\s+желез\w)[\s-]*([0123456]|[IV]{1,4})([^0123456IV]|$)",
               re.IGNORECASE)
replaceMol = r"[RADS\2]\3"

regexpFindRads2 = \
    re.compile(
        r"(?:BI(?:[-\s]{0,10})?)?(R?ADS)[^0123456IV]+прав[^0123456IV]+([0123456]|[IV]{1,2})[^0123456IV]+лев[^0123456IV]+([0123456]|[IV]{1,2})([^0123456IV]|$)",
        re.IGNORECASE)
replaceRads2 = r"[RADS\2], [RADS\3]\4"
regexpFindRads22 = \
    re.compile(
        r"(?:BI(?:[-\s]{0,10})?)?(R?ADS)[^0123456IV]+([0123456]|[IV]{1,2})[^0123456IV]+прав[^0123456IV]+([0123456]|[IV]{1,2})[^0123456IV]+лев[^0123456IV]{1,2}([^0123456IV]|$)",
        re.IGNORECASE)
replaceRads22 = r"[RADS\2], [RADS\3]\4"

regexpFindRadsError = \
    re.compile(r"(?:BI(?:[-\s]{0,10})?)?(R?ADS)[^0123456IV]{0,15}([0123456]{1,4})([^0123456IV]|$)", re.IGNORECASE)

# try:
#     filex = io.open(excelFile, 'wb')
#     filex.close()
# except io.BlockingIOError as ex:
#     print("")
#     print("ОШИБКА:: Закройте файл %s в Excel'e" % excelFile)
#     exit(33)
# try:
#     filex = io.open(excelFileOut, 'wb')
#     filex.close()
# except io.BlockingIOError as ex:
#     print("")
#     print("ОШИБКА:: Закройте файл %s в Excel'e" % excelFileOut,)
#     exit(33)


def get_rads(grads_list):
    grads_min = min(grads_list)
    grads_max = max(grads_list)
    # Количество радсов (2 - обычно, 1 - редко, 3 и более - странно)
    grads_count = len(grads_list)
    # Результат, большее или ноль
    grads = 0 if grads_min == 0 else grads_max

    return (grads, grads_count)


wb = load_workbook(filename=excelFile)
ws1 = wb.active
rowEnd = ws1.max_row

ws1["%s%s" % (columnPyResolution, 1)].value = "Все заключения"
ws1["%s%s" % (columnPyResolutionCheck, 1)].value = "Чистое заключение"
ws1["%s%s" % (columnPyRads, 1)].value = "RADS-A"
ws1["%s%s" % (columnPgmi, 1)].value = "PGMI-A"
ws1["%s%s" % (columnError, 1)].value = "Ошибка"
ws1["%s%s" % (columnPyResult, 1)].value = "Количество RADS-A"

ws1["%s%s" % (columnVerboseRads, 1)].value = "RADS-VA"
ws1["%s%s" % (columnVerbosePgmi, 1)].value = "PGMI-VA"
ws1["%s%s" % (columnKicker, 1)].value = "Kicker"

# print(ws1.max_row)

report = {}
report_line = {"org": "",
               "date": "",
               "rads-0": 0,
               "rads-1": 0,
               "rads-2": 0,
               "rads-3": 0,
               "rads-4": 0,
               "rads-5": 0,
               "rads-6": 0,
               "rads-54": 0,
               "pgmi-mi": 0,
               "rads-no": 0,
               }

byUids = {}

for row in range(rowStart, rowEnd):

    uid = ws1["%s%s" % (columnUid, row)].value
    try:
        date = datetime.strptime(ws1["%s%s" % (columnDateEdit, row)].value, '%d.%m.%Y %H:%M:%S')
    except:
        print(ws1["%s%s" % (columnDateEdit, row)].value)
        continue
    uid_date = date.strftime('%Y-%m-%d-%H-%M-%S')
    if uid not in byUids:
        byUids[uid] = {uid_date: row}
    else:
        byUids[uid][uid_date] = row

for uid in sorted(byUids):
    if len(byUids[uid]) < 2:
        del byUids[uid]


# Проходим по всем строка с данными
for row in range(rowStart, rowEnd):
    doc = ws1["%s%s" % (columnDoc, row)].value
    docExert = ws1["%s%s" % (columnDocExpert, row)].value

    # Строки с данными по мнению экселя не обязательно содержат данные необходимые нам
    if doc is not None:

        date = datetime.strptime(ws1["%s%s" % (columnDate, row)].value, '%d.%m.%Y %H:%M:%S')
        kicker_date = date.strftime('%Y.%m.%d')
        kicker = "%s---%s" % (kicker_date, ws1["%s%s" % (columnOrg, row)].value)
        ws1["%s%s" % (columnKicker, row)].value = kicker
        if kicker not in report:
            report[kicker] = {"org": ws1["%s%s" % (columnOrg, row)].value,
                              "date": date.strftime('%d.%m.%Y'),
                              "rads-0": 0,
                              "rads-1": 0,
                              "rads-2": 0,
                              "rads-3": 0,
                              "rads-4": 0,
                              "rads-5": 0,
                              "rads-6": 0,
                              "rads-54": 0,
                              "pgmi-mi": 0,
                              "rads-no": 0,
                              "SUM": 0,
                              }

        uid = ws1["%s%s" % (columnUid, row)].value
        if uid in byUids and byUids[uid][max(sorted(byUids[uid]))] != row:
            ws1["%s%s" % (columnError, row)].value = "Устарело (дубль)"
            ws1["%s%s" % (columnPyResult, row)].value = "OLD"
            ws1["%s%s" % (columnVerboseRads, row)].value = "OLD"
            ws1["%s%s" % (columnVerbosePgmi, row)].value = "OLD"

            # report[kicker]["rads-no"] = report[kicker]["rads-no"] + 1
            # report[kicker]["SUM"] = report[kicker]["SUM"] + 1

            continue

        resolution = ws1["%s%s" % (columnResolution, row)].value.replace("_x000D_", "").strip().replace(r"sin", "syn")
        print(">> ", row, doc, docExert)
        print("R ", resolution)

        # Пии участии эксперта нам важны только заключения эксперта
        if docExert is not None:
            pyResolutionCheckSplitted = re.split(
                r"(?:Дополнительное\s+заключение|ВТОРОЕ\s+ЧТЕНИЕ|Дополнительные сведения)", resolution, 1,
                re.IGNORECASE)

            if len(pyResolutionCheckSplitted) > 1:
                [_, resolution] = pyResolutionCheckSplitted

        # Жук любит вписывать в заключение рекомендации со всеми РАДСами
        # if re.match(r"(Жук\s+Александр\s+Романович|Бойцева\s+Елена\s+Анатольевна)", doc, re.IGNORECASE):
        pyResolutionCheckRecommendsSplit = re.split(r"Рекомендации:", resolution, 2, re.IGNORECASE)
        if len(pyResolutionCheckRecommendsSplit) == 2:
            [resolution, _] = pyResolutionCheckRecommendsSplit

        # double = False

        # Преобразуем заключение с помощью регулярки и заменяем римские цифры
        pyResolutionCheckMol = pyResolutionCheck = pyResolution = \
            regexpFindMol.sub(replaceMol, resolution) \
                .replace("[RADSVI]", "[RADS6]") \
                .replace("[RADSV]", "[RADS5]") \
                .replace("[RADSIV]", "[RADS4]") \
                .replace("[RADSIII]", "[RADS3]") \
                .replace("[RADSII]", "[RADS2]") \
                .replace("[RADSI]", "[RADS1]")
        # print(pyResolution)
        rads_list_mol = list(map(int, re.findall(r"\[RADS(\d)\]", pyResolutionCheck, re.IGNORECASE)))

        pyResolutionCheckRads = \
            regexpFindRads.sub(replaceRads, resolution) \
                .replace("[RADSVI]", "[RADS6]") \
                .replace("[RADSV]", "[RADS5]") \
                .replace("[RADSIV]", "[RADS4]") \
                .replace("[RADSIII]", "[RADS3]") \
                .replace("[RADSII]", "[RADS2]") \
                .replace("[RADSI]", "[RADS1]")

        if pyResolutionCheck.find("[RADS") == -1:
            pyResolutionCheck = pyResolution = pyResolutionCheckRads
            changed = True
        else:
            changed = False

        rads_list_rads = list(map(int, re.findall(r"\[RADS(\d)\]", pyResolutionCheckRads, re.IGNORECASE)))

        if pyResolutionCheckMol != resolution and pyResolutionCheckRads != resolution:
            double = True
            # ws1["%s%s" % (columnError, row)].value = "Двойные"
        else:
            double = False

        # Все радсы из очищенного заключения
        rads_list = list(map(int, re.findall(r"\[RADS(\d)\]", pyResolutionCheck, re.IGNORECASE)))
        print("RC ", pyResolutionCheck)

        if len(rads_list) == 1:
            pyResolutionCheck2 = \
                regexpFindRads2.sub(replaceRads2, resolution) \
                    .replace("[RADSVI]", "[RADS6]") \
                    .replace("[RADSV]", "[RADS5]") \
                    .replace("[RADSIV]", "[RADS4]") \
                    .replace("[RADSIII]", "[RADS3]") \
                    .replace("[RADSII]", "[RADS2]") \
                    .replace("[RADSI]", "[RADS1]")

            rads_list2 = list(map(int, re.findall(r"\[RADS(\d)\]", pyResolutionCheck2, re.IGNORECASE)))
            if len(rads_list2) > len(rads_list):
                pyResolutionCheck = pyResolutionCheck2
                rads_list = rads_list2

        if len(rads_list) == 1:
            pyResolutionCheck22 = \
                regexpFindRads22.sub(replaceRads22, resolution) \
                    .replace("[RADSVI]", "[RADS6]") \
                    .replace("[RADSV]", "[RADS5]") \
                    .replace("[RADSIV]", "[RADS4]") \
                    .replace("[RADSIII]", "[RADS3]") \
                    .replace("[RADSII]", "[RADS2]") \
                    .replace("[RADSI]", "[RADS1]")

            rads_list22 = list(map(int, re.findall(r"\[RADS(\d)\]", pyResolutionCheck22, re.IGNORECASE)))
            if len(rads_list22) > len(rads_list):
                pyResolutionCheck = pyResolutionCheck22
                rads_list = rads_list22

        ws1["%s%s" % (columnPyResolution, row)].value = pyResolution
        ws1["%s%s" % (columnPyResolutionCheck, row)].value = pyResolutionCheck


        if len(rads_list) == 0:
            ws1["%s%s" % (columnError, row)].value = "Нет значения"
            ws1["%s%s" % (columnPyResult, row)].value = ''
            ws1["%s%s" % (columnVerboseRads, row)].value = "NOT FOUND"
            ws1["%s%s" % (columnVerbosePgmi, row)].value = "NO DATA"

            report[kicker]["rads-no"] = report[kicker]["rads-no"] + 1
            report[kicker]["SUM"] = report[kicker]["SUM"] + 1

            continue

        if len(list(map(int, re.findall(r"R?ADS[^0123456IV]{0,15}([0123456]{2,4})(?:[^0123456IV]|$)", resolution,
                                        re.IGNORECASE)))) > 0:
            ws1["%s%s" % (columnError, row)].value = "Неверное"
            ws1["%s%s" % (columnVerboseRads, row)].value = "NOT FOUND"
            ws1["%s%s" % (columnPyResult, row)].value = ''
            ws1["%s%s" % (columnVerbosePgmi, row)].value = "NO DATA"

            continue


        (rads, rads_count) = get_rads(rads_list)
        if double:
            # print("CCC ", rads_list_mol, rads_list)
            ws1["%s%s" % (columnErrorComment, row)].value = "%s%s" % (rads_list_mol, rads_list_rads)
            if rads_list_mol == rads_list_rads:
                ws1["%s%s" % (columnError, row)].value = "Несколько, но совпадает"
            else:
                (radsM, rads_countM) = get_rads(rads_list_mol)
                (radsR, rads_countR) = get_rads(rads_list_rads)

                if radsR == radsM:
                    ws1["%s%s" % (columnError, row)].value = "Несколько, результат один"
                else:
                    ws1["%s%s" % (columnError, row)].value = "Несколько!"
                    ws1["%s%s" % (columnVerboseRads, row)].value = "NOT FOUND"
                    ws1["%s%s" % (columnPyResult, row)].value = ''

                    continue


        #print("RADS ", rads, rads_count, rads_list, rads_min, rads_max)

        # Заполняем данные
        ws1["%s%s" % (columnPyRads, row)].value = rads
        ws1["%s%s" % (columnVerboseRads, row)].value = rads

        if ws1["%s%s" % (columnDescription, row)].value is not None:
            pgmi = list(map(str, regexpFindPgmi.findall(ws1["%s%s" % (columnDescription, row)].value.replace("\n", ""))))
            if len(pgmi) > 0:
                ws1["%s%s" % (columnPgmi, row)].value = pgmi[0].upper().replace("Р", "P").replace("М", "M")
                pgmi_data = pgmi[0].upper().replace("Р", "P").replace("М", "M")
                ws1["%s%s" % (columnVerbosePgmi, row)].value = "PGMI: %s" % pgmi_data

                if pgmi_data == "I" or pgmi_data == "M":
                    report[kicker]["pgmi-mi"] = report[kicker]["pgmi-mi"] + 1
            else:
                ws1["%s%s" % (columnVerbosePgmi, row)].value = "NO DATA"
        else:
            ws1["%s%s" % (columnVerbosePgmi, row)].value = "NO DATA"

        ws1["%s%s" % (columnPyResult, row)].value = rads_count

        report[kicker]["rads-%d" % rads] = report[kicker]["rads-%d" % rads] + 1
        report[kicker]["SUM"] = report[kicker]["SUM"] + 1
        if rads == 5 or rads == 4:
            report[kicker]["rads-54"] = report[kicker]["rads-54"] + 1


ws1.auto_filter.ref = "A1:AF%d" % rowEnd
ws1.freeze_panes = ws1['A2']

ws2 = wb.create_sheet(title="Report")

#print(report)

rrow = 2
rdate = None
rdate1 = 0
fontbold = Font(bold=True)

allSum = []

alignmentH = Alignment(horizontal='center', vertical='center', wrap_text=True)
alignmentO = Alignment(shrink_to_fit=True)

for ii in range(0,25):
    ws2["%s%s" % (chr(ord('A')+ii), 1)].alignment = alignmentH
    ws2["%s%s" % (chr(ord('A')+ii), 1)].font = fontbold


ws2["%s%s" % ("A", 1)].value = 'МО'
ws2["%s%s" % ("B", 1)].value = 'Дата проведения исследований'
ws2["%s%s" % ("D", 1)].value = 'Кол-во ММГ исследований ЕРИС'
ws2["%s%s" % ("F", 1)].value = 'Количество BI-RADS: 0'
ws2["%s%s" % ("G", 1)].value = '% BI-RADS: 0 от числа всех ММГ'
ws2["%s%s" % ("H", 1)].value = 'Количество BI-RADS: 1'
ws2["%s%s" % ("I", 1)].value = '% BI-RADS: 1 от числа всех ММГ'
ws2["%s%s" % ("J", 1)].value = 'Количество BI-RADS: 2'
ws2["%s%s" % ("K", 1)].value = '% BI-RADS: 2 от числа всех ММГ'
ws2["%s%s" % ("L", 1)].value = 'Количество BI-RADS: 3'
ws2["%s%s" % ("M", 1)].value = '% BI-RADS: 3 от числа всех ММГ'
ws2["%s%s" % ("N", 1)].value = 'Количество BI-RADS: 4'
ws2["%s%s" % ("O", 1)].value = '% BI-RADS: 4 от числа всех ММГ'
ws2["%s%s" % ("P", 1)].value = 'Количество BI-RADS: 5'
ws2["%s%s" % ("Q", 1)].value = '% BI-RADS: 5 от числа всех ММГ'
ws2["%s%s" % ("R", 1)].value = 'Количество исследований с указанием BI-RADS: 4-5'
ws2["%s%s" % ("S", 1)].value = '% BI-RADS: 4-5 от числа всех ММГ'
ws2["%s%s" % ("T", 1)].value = 'Количество BI-RADS: 6'
ws2["%s%s" % ("U", 1)].value = '% BI-RADS: 6 от числа всех ММГ'
ws2["%s%s" % ("V", 1)].value = 'Количество ОШИБОЧНЫХ ЗАКЛЮЧЕНИЙ в BI-RADS'
ws2["%s%s" % ("W", 1)].value = '% ОШИБОЧНЫХ ЗАКЛЮЧЕНИЙ в BI-RADS от числа всех ММГ'
ws2["%s%s" % ("X", 1)].value = 'Количество M и I степеней по системе PGMI'
ws2["%s%s" % ("Y", 1)].value = 'Доля выбранных M и I степеней от числа всех проведенных ММГ'


def sum_row(srrow, srdate, srdate1):

    global allSum

    # ws2["%s%s" % ("A", srrow)].value = ''
    ws2["%s%s" % ("B", srrow)].value = srdate
    # ws2["%s%s" % ("A", srrow):"%s%s" % ("V", rrow)].font = fontbold
    # ws2["%s%s" % ("A", srrow):"%s%s" % ("V", rrow)].font = fontbold
    ws2["%s%s" % ("B", srrow)].font = fontbold
    ws2["%s%s" % ("D", srrow)].font = fontbold
    ws2["%s%s" % ("E", srrow)].font = fontbold

    ws2["%s%s" % ("D", srrow)].value = "=SUM(D%d:D%d)" % (srdate1, srrow - 1)
    ws2["%s%s" % ("E", srrow)].value = "=F%d+H%d+J%d+L%d+N%d+P%d+T%d+V%d" % (
        srrow, srrow, srrow, srrow, srrow, srrow, srrow, srrow)

    for ii in range(0, 10):
        iil = chr(ord('F') + (ii * 2))
        iil2 = chr(ord('F') + (ii * 2) + 1)

        ws2["%s%s" % (iil, srrow)].value = "=SUM(%s%d:%s%d)" % (iil, srdate1, iil, srrow - 1)
        ws2["%s%s" % (iil2, srrow)].value = "=%s%d/D%d" % (iil, srrow, srrow)
        ws2["%s%s" % (iil, srrow)].font = fontbold
        ws2["%s%s" % (iil2, srrow)].number_format = FORMAT_PERCENTAGE_00
        ws2["%s%s" % (iil2, srrow)].font = fontbold

    allSum.append(str(srrow))


for report_order in sorted(report):

    if rdate is None:
        rdate = report[report_order]['date']
        rdate1 = rrow

    if rdate != report[report_order]['date']:
        sum_row(rrow, rdate, rdate1)
        # #ws2["%s%s" % ("A", rrow)].value = ''
        # ws2["%s%s" % ("B", rrow)].value = rdate
        # #ws2["%s%s" % ("A", rrow):"%s%s" % ("V", rrow)].font = fontbold
        # #ws2["%s%s" % ("A", rrow):"%s%s" % ("V", rrow)].font = fontbold
        # ws2["%s%s" % ("B", rrow)].font = fontbold
        # ws2["%s%s" % ("D", rrow)].font = fontbold
        # ws2["%s%s" % ("E", rrow)].font = fontbold
        #
        # ws2["%s%s" % ("D", rrow)].value = "=SUM(D%d:D%d)" % (rdate1, rrow-1)
        # ws2["%s%s" % ("E", rrow)].value = "=F%d+H%d+J%d+L%d+N%d+P%d+T%d+V%d" % (
        #     rrow, rrow, rrow, rrow, rrow, rrow, rrow, rrow)
        #
        # for ii in range(0, 10):
        #     iil = chr(ord('F') + (ii * 2))
        #     iil2 = chr(ord('F') + (ii * 2) + 1)
        #
        #     ws2["%s%s" % (iil, rrow)].value = "=SUM(%s%d:%s%d)" % (iil, rdate1, iil, rrow - 1)
        #     ws2["%s%s" % (iil2, rrow)].value = "=%s%d/D%d" % (iil, rrow, rrow)
        #     ws2["%s%s" % (iil, rrow)].font = fontbold
        #     ws2["%s%s" % (iil2, rrow)].number_format = FORMAT_PERCENTAGE_00
        #     ws2["%s%s" % (iil2, rrow)].font = fontbold

        # allSum.append(str(rrow))
        rrow = rrow + 1
        rdate = report[report_order]['date']
        rdate1 = rrow

    ws2["%s%s" % ("A", rrow)].value = report[report_order]['org']
    ws2["%s%s" % ("A", rrow)].alignment = alignmentO

    orgWidth = (len(str(ws2["%s%s" % ("A", rrow)].value)) + 2) * 1.2
    if orgWidth > ws2.column_dimensions["A"].width:
        ws2.column_dimensions["A"].width = orgWidth

    ws2["%s%s" % ("B", rrow)].value = report[report_order]['date']
    ws2["%s%s" % ("B", rrow)].alignment = alignmentO

    orgWidth = (len(str(ws2["%s%s" % ("B", rrow)].value)) + 2) * 1.2
    if orgWidth > ws2.column_dimensions["B"].width:
        ws2.column_dimensions["B"].width = orgWidth

    ws2["%s%s" % ("D", rrow)].value = report[report_order]['SUM']
    ws2["%s%s" % ("E", rrow)].value = "=F%d+H%d+J%d+L%d+N%d+P%d+T%d+V%d" % (
        rrow, rrow, rrow, rrow, rrow, rrow, rrow, rrow)
    ws2["%s%s" % ("F", rrow)].value = report[report_order]['rads-0']
    ws2["%s%s" % ("G", rrow)].value = "=F%d/D%d" % (rrow, rrow)
    ws2["%s%s" % ("G", rrow)].number_format = FORMAT_PERCENTAGE_00
    ws2["%s%s" % ("H", rrow)].value = report[report_order]['rads-1']
    ws2["%s%s" % ("I", rrow)].value = "=H%d/D%d" % (rrow, rrow)
    ws2["%s%s" % ("I", rrow)].number_format = FORMAT_PERCENTAGE_00
    ws2["%s%s" % ("J", rrow)].value = report[report_order]['rads-2']
    ws2["%s%s" % ("K", rrow)].value = "=J%d/D%d" % (rrow, rrow)
    ws2["%s%s" % ("K", rrow)].number_format = FORMAT_PERCENTAGE_00
    ws2["%s%s" % ("L", rrow)].value = report[report_order]['rads-3']
    ws2["%s%s" % ("M", rrow)].value = "=L%d/D%d" % (rrow, rrow)
    ws2["%s%s" % ("M", rrow)].number_format = FORMAT_PERCENTAGE_00
    ws2["%s%s" % ("N", rrow)].value = report[report_order]['rads-4']
    ws2["%s%s" % ("O", rrow)].value = "=N%d/D%d" % (rrow, rrow)
    ws2["%s%s" % ("O", rrow)].number_format = FORMAT_PERCENTAGE_00
    ws2["%s%s" % ("P", rrow)].value = report[report_order]['rads-5']
    ws2["%s%s" % ("Q", rrow)].value = "=P%d/D%d" % (rrow, rrow)
    ws2["%s%s" % ("Q", rrow)].number_format = FORMAT_PERCENTAGE_00
    ws2["%s%s" % ("R", rrow)].value = report[report_order]['rads-54']
    ws2["%s%s" % ("S", rrow)].value = "=R%d/D%d" % (rrow, rrow)
    ws2["%s%s" % ("S", rrow)].number_format = FORMAT_PERCENTAGE_00
    ws2["%s%s" % ("T", rrow)].value = report[report_order]['rads-6']
    ws2["%s%s" % ("U", rrow)].value = "=T%d/D%d" % (rrow, rrow)
    ws2["%s%s" % ("U", rrow)].number_format = FORMAT_PERCENTAGE_00
    ws2["%s%s" % ("V", rrow)].value = report[report_order]['rads-no']
    ws2["%s%s" % ("W", rrow)].value = "=V%d/D%d" % (rrow, rrow)
    ws2["%s%s" % ("W", rrow)].number_format = FORMAT_PERCENTAGE_00
    ws2["%s%s" % ("X", rrow)].value = report[report_order]['pgmi-mi']
    ws2["%s%s" % ("Y", rrow)].value = "=X%d/D%d" % (rrow, rrow)
    ws2["%s%s" % ("Y", rrow)].number_format = FORMAT_PERCENTAGE_00

    rrow = rrow + 1

sum_row(rrow, rdate, rdate1)

rrow = rrow + 2
ws2["%s%s" % ("B", rrow)].value = "ИТОГ"
ws2["%s%s" % ("B", rrow)].font = fontbold
for ii in range(0, 11):
    iil = chr(ord('D') + (ii * 2))
    iil2 = chr(ord('D') + (ii * 2) + 1)

    allFormula = "=%s0%s" %(iil, ("+%s" % iil).join(allSum))
    ws2["%s%s" % (iil, rrow)].value = allFormula
    ws2["%s%s" % (iil, rrow)].font = fontbold
    if iil2 == "E":
        allFormula = "=%s0%s" %(iil2, ("+%s" % iil2).join(allSum))
        ws2["%s%s" % (iil2, rrow)].value = allFormula
        continue
    ws2["%s%s" % (iil2, rrow)].value = "=%s%d/D%d" % (iil, rrow, rrow)
    ws2["%s%s" % (iil, rrow)].font = fontbold
    ws2["%s%s" % (iil2, rrow)].number_format = FORMAT_PERCENTAGE_00
    ws2["%s%s" % (iil2, rrow)].font = fontbold

ws2.freeze_panes = ws1['C2']


# Сохраняем в "соседнем" файле
wb.save(filename=excelFileOut)
